import openpyxl as xl
from openpyxl.chart import BarChart, Reference, PieChart

def workOnSheet(sheetName, page, valuePos, correctPos):
    workBook = xl.load_workbook(sheetName)
    sheet = workBook[page]

    for row in range(2, sheet.max_row + 1):
        cellValue = sheet.cell(row, valuePos)
        corrected = cellValue.value * 0.9
        correctedCell = sheet.cell(row, correctPos)
        correctedCell.value = corrected

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart1 = PieChart()

    chart.add_data(values)
    chart1.add_data(values)

    sheet.add_chart(chart, 'g2')
    sheet.add_chart(chart1, 'a8')

    workBook.save(sheetName)

workOnSheet("items.xlsx", "Sheet1", 3, 5)


# workBook = xl.load_workbook("items.xlsx")
# sheet = workBook['Sheet1']
# # cell = sheet['a2']
# # print(cell.value)
# # print(sheet.max_row)

# for row in range(2, sheet.max_row + 1):
#     cellValue = sheet.cell(row, 3)
#     corrected = cellValue.value * 0.9
#     correctedCell = sheet.cell(row, 5)
#     correctedCell.value = corrected

# values = Reference(sheet, max_row=sheet.max_row, min_row=2, min_col=4, max_col=4) 
# chart = BarChart()
# chart1 = PieChart()

# chart.add_data(values)
# chart1.add_data(values)

# sheet.add_chart(chart, "f2")  
# sheet.add_chart(chart1, "a8") 

# workBook.save('items2.xlsx')    